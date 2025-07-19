import os
import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.filters import CommandStart, Text
from openpyxl import load_workbook
from datetime import datetime
import tempfile
import requests

# Конфигурация
TOKEN = os.environ.get("TOKEN")
if not TOKEN:
    raise ValueError("Токен не найден. Установите переменную окружения TOKEN.")

USER_ID = 4272684  # Искомый номер пользователя

PROGRAMS = {
    "hse": {
        "name": "📊 Экономика",
        "url": "https://enrol.hse.ru/storage/public_report_2025/moscow/Bachelors/BD_moscow_Economy_O.xlsx",
        "priority": 1,
        "places": 10
    },
    "resh": {
        "name": "📘 Совбак НИУ ВШЭ и РЭШ",
        "url": "https://enrol.hse.ru/storage/public_report_2025/moscow/Bachelors/BD_moscow_RESH_O.xlsx",
        "priority": 2,
        "places": 6
    }
}

# Инициализация бота
bot = Bot(token=TOKEN)
dp = Dispatcher()

# Клавиатура
keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=PROGRAMS["hse"]["name"])],
        [KeyboardButton(text=PROGRAMS["resh"]["name"])]
    ],
    resize_keyboard=True
)

@dp.message(CommandStart())
async def start(message: types.Message):
    await message.answer("Выбери программу:", reply_markup=keyboard)

@dp.message(Text(text=[p["name"] for p in PROGRAMS.values()]))
async def handle_selection(message: types.Message):
    for code, data in PROGRAMS.items():
        if message.text == data["name"]:
            text = await process_program(code)
            await message.answer(text)
            break

async def process_program(code):
    data = PROGRAMS[code]
    url = data["url"]
    priority = data["priority"]
    places = data["places"]

    # Скачиваем файл
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        response = requests.get(url)
        tmp.write(response.content)
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)
    ws = wb.active

    user_score = None
    competitors = []
    same_priority_rank = 0

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            number = int(row[1].value)
            agree = row[9].value
            row_priority = int(row[11].value)
            score = float(row[18].value)
        except (TypeError, ValueError):
            continue

        if agree != "Да":
            continue

        if number == USER_ID:
            user_score = score

        # Рейтинг по приоритету
        if row_priority == priority:
            competitors.append(score)

    if user_score is None:
        return "❗ Не найден пользователь с указанным номером."

    # Считаем рейтинг по своему приоритету
    sorted_scores = sorted([s for s in competitors if s > 0], reverse=True)
    same_priority_rank = sum(s > user_score for s in sorted_scores) + 1

    # Считаем конкурентов с более высоким баллом на другом приоритете
    alt_priority = 3 - priority
    higher_alt = 0
    for row in ws.iter_rows(min_row=2):
        try:
            number = int(row[1].value)
            agree = row[9].value
            row_priority = int(row[11].value)
            score = float(row[18].value)
        except (TypeError, ValueError):
            continue

        if agree != "Да":
            continue

        if row_priority == alt_priority and score > user_score:
            higher_alt += 1

    now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    return (
        f"📅 Дата обновления данных: {now}\n\n"
        f"🎯 Мест на программе: {places}\n\n"
        f"✅ Твой рейтинг среди {priority} приоритета: {same_priority_rank}\n\n"
        f"🔺 Людей с {alt_priority} приоритетом и баллом выше: {higher_alt}"
    )


async def main():
    logging.basicConfig(level=logging.INFO)
    await dp.start_polling(bot)

if __name__ == "__main__":
    import asyncio
    asyncio.get_event_loop().run_until_complete(main())
